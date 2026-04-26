"""
Modul batch -- Halaman analisis batch (ZIP).
"""

import io
import os
import threading

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go

from config import DEFAULT_SUBBANDS, DEFAULT_FEATURES, TASK_COLORS, ACCENT_LIGHT
from processing.features import EEGFeatures

from processing.loader import EEGLoader
from processing.filters import EEGFilters
from processing.features import EEGFeatures
from processing.delta import DeltaCalculator
from processing.epoching import EpochEngine

from visualization.feature_plots import FeaturePlots
from visualization.comparison_plots import ComparisonPlots


# ---------------------------------------------------------------------------
# Worker: proses satu file EDF
# ---------------------------------------------------------------------------

def _process_single_edf(zip_source, edf_path, channels, subbands, features,
                         include_frequency=False,
                         use_amplitude=False, use_ica=False,
                         ica_n=None, ica_method="fastica",
                         psd_method="welch", psd_fmin=0.0,
                         psd_fmax=49.0, psd_n_fft=None,
                         use_epoching=False, epoch_mode="fixed",
                         epoch_duration=2.0, window_overlap=0.5,
                         use_epoch_reject=False, epoch_reject_threshold=100.0,
                         use_connectivity=False, connectivity_method="wpli",
                         connectivity_channels=None,
                         cache_dir=None):
    """Worker: proses satu file EDF dari ZIP bytes (thread-safe).

    Jika ``cache_dir`` diberikan, DataFrame preprocessed disimpan ke
    pickle di dalam folder tersebut. Path cache + metadata dikembalikan
    pada field ``cache_info`` sehingga step encoding post-batch bisa
    langsung pakai data ini tanpa load EDF ulang.

    Returns
    -------
    tuple  (feat_df, tasks_found, erd_df, conn_df, cache_info)
           feat_df    : fitur per task/channel/subband (atau None)
           tasks_found: list nama task
           erd_df     : DataFrame ERD/ERS paired per task (atau None)
           conn_df    : DataFrame connectivity (atau None)
           cache_info : dict {cache_path, filename, subject_id,
                              scenario, scenario_id, channels, sfreq,
                              tasks} atau None
    """
    from processing.connectivity import ConnectivityAnalyzer
    meta = EEGLoader.detect_category(edf_path)
    loader = EEGLoader()

    try:
        if isinstance(zip_source, (str, bytes, os.PathLike)) and os.path.isfile(zip_source):
            with open(zip_source, "rb") as f:
                loader.load_edf_from_zip(f, edf_path)
        else:
            buf = io.BytesIO(zip_source)
            loader.load_edf_from_zip(buf, edf_path)
    except Exception as e:
        print(f"Worker file error ({edf_path}): {e}")
        return None, [], None, None, None

    ch_list = channels if channels else loader.channel_names
    ch_list = [c for c in ch_list if c in loader.channel_names]
    if not ch_list:
        loader._cleanup_tmp()
        return None, [], None, None, None

    # Amplitude filter (clipping artefak)
    if use_amplitude:
        try:
            EEGFilters.apply_amplitude_filter(loader)
        except Exception:
            pass

    try:
        low_all = min(v[0] for v in subbands.values())
        high_all = max(v[1] for v in subbands.values())
        EEGFilters.apply_bandpass(loader, low_all, high_all)
    except Exception:
        pass

    # ICA (hapus artefak otomatis)
    if use_ica:
        try:
            EEGFilters.apply_ica(loader, n_components=ica_n, method=ica_method)
        except Exception:
            pass

    df = loader.extract_dataframe()
    tasks_found = [t for t in loader.get_task_list() if t != "none"]

    # --- Cache preprocessed DataFrame untuk post-batch encoding ---
    cache_info = None
    if cache_dir is not None:
        try:
            # Parse subject_id dan scenario_id dari meta (struktur EEGET-ALS)
            subject_id = meta.get("subject", "unknown")
            scenario_str = meta.get("scenario", "")  # e.g. "scenario5"
            try:
                scenario_id = int(scenario_str.replace("scenario", ""))
            except (ValueError, AttributeError):
                scenario_id = -1

            safe_name = (
                edf_path.replace("/", "__")
                        .replace("\\", "__")
                        .replace(" ", "_")
            )
            cache_path = os.path.join(cache_dir, f"{safe_name}.pkl")
            df.to_pickle(cache_path)

            cache_info = {
                "cache_path": cache_path,
                "filename": edf_path,
                "subject_id": subject_id,
                "scenario": scenario_str if scenario_str else "unknown",
                "scenario_id": scenario_id,
                "channels": list(loader.channel_names),
                "sfreq": float(loader.sfreq),
                "tasks": list(tasks_found),
            }
        except Exception:
            cache_info = None

    if not tasks_found:
        loader._cleanup_tmp()
        return None, tasks_found, None, None, cache_info

    if use_epoching:
        if epoch_mode == "fixed":
            feat_df = EpochEngine.compute_epoched_task_features(
                loader, df, ch_list, tasks_found, loader.sfreq, subbands, features,
                epoch_duration=epoch_duration,
                reject_threshold=epoch_reject_threshold if use_epoch_reject else None,
                include_frequency=include_frequency,
                psd_method=psd_method, psd_fmin=psd_fmin,
                psd_fmax=psd_fmax, psd_n_fft=psd_n_fft,
            )
            # drop n_epochs etc or keep them
        else:
            win_df = EpochEngine.compute_windowed_task_features(
                loader, df, ch_list, tasks_found, loader.sfreq, subbands, features,
                window_size=epoch_duration, overlap=window_overlap,
                include_frequency=include_frequency,
                psd_method=psd_method, psd_fmin=psd_fmin,
                psd_fmax=psd_fmax, psd_n_fft=psd_n_fft,
            )
            feat_df = EpochEngine.aggregate_windowed_features(win_df)
    else:
        feat_df = EEGFeatures.compute_first_occurrence_features(
        loader, df, ch_list, tasks_found, subbands, features,
        include_frequency=include_frequency,
        psd_method=psd_method, psd_fmin=psd_fmin,
        psd_fmax=psd_fmax, psd_n_fft=psd_n_fft,
    )

    # --- Hitung ERD/ERS paired (Resting → Task) per non-Resting task ---
    erd_parts = []
    non_resting_tasks = [t for t in tasks_found if t != "Resting"]
    if "Resting" in tasks_found and non_resting_tasks:
        for task_name in non_resting_tasks:
            erd_one = EEGFeatures.compute_erd_ers_paired(
                loader, df, ch_list, task_name,
                subbands=subbands,
                baseline_task="Resting",
            )
            if not erd_one.empty:
                erd_parts.append(erd_one)

    loader._cleanup_tmp()

    # Gabungkan ERD/ERS dari semua task
    erd_df = None
    if erd_parts:
        erd_df = pd.concat(erd_parts, ignore_index=True)
        erd_df.insert(0, "filename", edf_path)
        erd_df.insert(1, "category", meta["category"])
        erd_df.insert(2, "subject", meta["subject"])
        erd_df.insert(3, "time", meta["time"])
        erd_df.insert(4, "scenario", meta["scenario"])

    # --- Hitung Connectivity (BARU) ---
    conn_df = None
    if use_connectivity:
        ch_list_conn = ch_list
        if connectivity_channels:
            ch_list_conn = [c for c in ch_list if c in connectivity_channels]
            if not ch_list_conn:
                ch_list_conn = ch_list
                
        conn_dict = ConnectivityAnalyzer.compute_task_connectivity(
            loader, df, ch_list_conn, tasks_found, loader.sfreq,
            method=connectivity_method, subbands=subbands,
            use_epoching=use_epoching, epoch_duration=epoch_duration,
        )
        if conn_dict:
            conn_df = ConnectivityAnalyzer.all_tasks_to_dataframe(
                conn_dict, ch_list_conn, method=connectivity_method
            )
            if not conn_df.empty:
                # Tambahkan meta kolom
                conn_df.insert(0, "filename", edf_path)
                conn_df.insert(1, "category", meta["category"])
                conn_df.insert(2, "subject", meta["subject"])
                conn_df.insert(3, "time", meta["time"])
                conn_df.insert(4, "scenario", meta["scenario"])

    if feat_df.empty:
        return None, tasks_found, erd_df, conn_df, cache_info

    feat_df.insert(0, "filename", edf_path)
    feat_df.insert(1, "category", meta["category"])
    feat_df.insert(2, "subject", meta["subject"])
    feat_df.insert(3, "time", meta["time"])
    feat_df.insert(4, "scenario", meta["scenario"])

    return feat_df, tasks_found, erd_df, conn_df, cache_info


# ---------------------------------------------------------------------------
# Batch processing
# ---------------------------------------------------------------------------

def run_batch_processing(cfg):
    """Jalankan batch analysis: baca semua EDF dari ZIP, hitung fitur.

    Selain menghitung fitur per-task, juga men-cache DataFrame
    preprocessed tiap file ke pickle di dalam folder temp. Cache ini
    dipakai oleh step post-batch encoding (lihat
    ``ui/encoding.py::render_post_batch_encoding``) sehingga tidak
    perlu load EDF ulang.
    """
    import shutil
    import tempfile
    from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed

    uploaded = cfg.get("uploaded")
    if uploaded is None:
        st.warning("File ZIP belum diunggah.")
        return

    # Bersihkan cache sebelumnya kalau ada
    prev_cache_dir = st.session_state.get("batch_cache_dir")
    if prev_cache_dir and os.path.isdir(prev_cache_dir):
        try:
            shutil.rmtree(prev_cache_dir, ignore_errors=True)
        except Exception:
            pass
    cache_dir = tempfile.mkdtemp(prefix="eeg_batch_cache_")
    st.session_state.batch_cache_dir = cache_dir
    st.session_state.batch_cached_files = []

    subbands = cfg["subbands"] or DEFAULT_SUBBANDS
    features = cfg["features"] or DEFAULT_FEATURES
    channels = cfg["channels"] if cfg["channels"] else None
    include_freq = cfg.get("include_frequency", False)
    use_amplitude = cfg.get("use_amplitude", False)
    use_ica = cfg.get("use_ica", False)
    ica_n = cfg.get("ica_n", None)
    ica_method = cfg.get("ica_method", "fastica")
    psd_method = cfg.get("psd_method", "welch")
    psd_fmin = cfg.get("psd_fmin", 0.0)
    psd_fmax = cfg.get("psd_fmax", 49.0)
    psd_n_fft = cfg.get("psd_n_fft", None)
    
    use_epoching = cfg.get("use_epoching", False)
    epoch_mode = cfg.get("epoch_mode", "fixed")
    epoch_duration = cfg.get("epoch_duration", 2.0)
    window_overlap = cfg.get("window_overlap", 0.5)
    use_epoch_reject = cfg.get("use_epoch_reject", False)
    epoch_reject_threshold = cfg.get("epoch_reject_threshold", 100.0)
    
    use_connectivity = cfg.get("use_connectivity", False)
    connectivity_method = cfg.get("connectivity_method", "wpli")
    connectivity_channels = cfg.get("connectivity_channels", [])
    progress_bar = st.progress(0, text="Memulai batch processing...")

    uploaded.seek(0)
    zip_bytes = uploaded.read()
    uploaded.seek(0)
    edf_list = EEGLoader.list_edf_in_zip(uploaded)

    if not edf_list:
        st.error("Tidak ditemukan file EDF dalam ZIP.")
        return

    # Tulis zip_bytes ke temporary file untuk mencegah overhead IPC
    # yang menyebabkan BrokenProcessPool di Windows.
    temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    temp_zip.write(zip_bytes)
    temp_zip.close()
    temp_zip_path = temp_zip.name

    all_dfs = []
    all_erd_dfs = []
    all_conn_dfs = []
    all_cache_infos = []
    common_tasks = set()
    n_total = len(edf_list)

    cpu_count = max(1, os.cpu_count() or 1)
    max_workers = min(cpu_count, n_total)

    # ProcessPoolExecutor untuk CPU-bound (MNE/NumPy melewati GIL).
    # Fallback ke ThreadPoolExecutor jika spawn gagal (misalnya di environment terbatas).
    PoolExecutor = ProcessPoolExecutor
    try:
        with ProcessPoolExecutor(max_workers=1) as _test:
            _test.submit(int, 0).result()
    except Exception:
        PoolExecutor = ThreadPoolExecutor

    try:
        with PoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(
                    _process_single_edf, temp_zip_path, path,
                    channels, subbands, features, include_freq,
                    use_amplitude, use_ica, ica_n, ica_method,
                    psd_method, psd_fmin, psd_fmax, psd_n_fft,
                    use_epoching, epoch_mode, epoch_duration,
                    window_overlap, use_epoch_reject, epoch_reject_threshold,
                    use_connectivity, connectivity_method, connectivity_channels,
                    cache_dir,
                ): path
                for path in edf_list
            }

            for i, future in enumerate(as_completed(futures), 1):
                path = futures[future]
                progress_bar.progress(
                    i / n_total,
                    text=f"Memproses {i}/{n_total}: {path.split('/')[-1]}",
                )
                try:
                    result = future.result()
                    if result is None:
                        continue
                    feat_df, tasks_found, erd_df, conn_df, cache_info = result
                    if feat_df is not None and not feat_df.empty:
                        all_dfs.append(feat_df)
                        if not common_tasks:
                            common_tasks = set(tasks_found)
                        else:
                            common_tasks |= set(tasks_found)
                    if erd_df is not None and not erd_df.empty:
                        all_erd_dfs.append(erd_df)
                    if conn_df is not None and not conn_df.empty:
                        all_conn_dfs.append(conn_df)
                    if cache_info is not None:
                        all_cache_infos.append(cache_info)
                except Exception as e:
                    print(f"Worker future exception on {path}: {e}")
                    pass
    finally:
        try:
            os.unlink(temp_zip_path)
        except OSError:
            pass

    progress_bar.empty()

    if not all_dfs:
        st.error("Tidak ada data fitur yang berhasil diekstrak.")
        return

    batch_df = pd.concat(all_dfs, ignore_index=True)
    if common_tasks is None:
        common_tasks = set()

    # Simpan ERD/ERS paired results
    batch_erd_df = pd.DataFrame()
    if all_erd_dfs:
        batch_erd_df = pd.concat(all_erd_dfs, ignore_index=True)

    # Simpan Connectivity results
    batch_conn_df = pd.DataFrame()
    if all_conn_dfs:
        batch_conn_df = pd.concat(all_conn_dfs, ignore_index=True)

    st.session_state.batch_df = batch_df
    st.session_state.batch_erd_df = batch_erd_df
    st.session_state.batch_conn_df = batch_conn_df
    st.session_state.batch_tasks = sorted(common_tasks)
    st.session_state.batch_cached_files = all_cache_infos
    st.session_state.batch_processed = True

    # Reset hasil encoding post-batch sebelumnya
    for key in (
        "encoding_done", "encoding_result", "encoding_summary",
        "encoding_elapsed", "encoding_output_path", "encoding_raw",
        "chunking_done", "chunking_features_df", "chunking_chain_df",
        "chunking_summary_df", "chunking_run_summary",
        "chunking_elapsed", "chunking_paths",
        # Comparison state
        "comparison_done", "comparison_summary_df",
        "comparison_details_df", "comparison_elapsed", "comparison_meta",
    ):
        st.session_state.pop(key, None)

    st.success(
        f"Batch processing selesai: {batch_df['filename'].nunique()} file, "
        f"{len(common_tasks)} task ditemukan. "
        f"{len(all_cache_infos)} file ter-cache untuk encoding."
    )


# ---------------------------------------------------------------------------
# Render hasil batch
# ---------------------------------------------------------------------------

def render_batch_results(cfg):
    """Tampilkan hasil batch analysis & delta comparison."""
    from ui.encoding import render_post_batch_encoding

    batch_df = st.session_state.batch_df
    batch_tasks = st.session_state.batch_tasks

    if batch_df is None or batch_df.empty:
        return

    # Kumpulkan opsi filter
    categories = sorted(batch_df["category"].unique().tolist()) if "category" in batch_df.columns else []
    scenarios = sorted(batch_df["scenario"].unique().tolist()) if "scenario" in batch_df.columns else []
    times = sorted(batch_df["time"].unique().tolist()) if "time" in batch_df.columns else []
    all_subbands = sorted(batch_df["subband"].unique().tolist()) if "subband" in batch_df.columns else []
    all_channels = sorted(batch_df["channel"].unique().tolist()) if "channel" in batch_df.columns else []

    # Ringkasan Batch (selalu ditampilkan di atas)
    st.markdown('<p class="section-title">Ringkasan Batch</p>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Kategori", categories[0] if len(categories) == 1 else f"{len(categories)} kategori")
    c2.metric("Subjek", batch_df["subject"].nunique() if "subject" in batch_df.columns else "-")
    c3.metric("Task Ditemukan", len(batch_tasks))
    c4.metric("Total Baris", f"{len(batch_df):,}")

    # Task badges
    badge_html = ""
    for t in batch_tasks:
        color = TASK_COLORS.get(t, ACCENT_LIGHT)
        badge_html += f'<span class="task-badge" style="background:{color}">{t}</span>'
    st.markdown(f'<div style="margin:8px 0">{badge_html}</div>', unsafe_allow_html=True)

    # ================================================================ #
    #  Tab Navigation: Delta | Encoding                                #
    # ================================================================ #
    tab_delta, tab_encoding = st.tabs([
        "📊 Perbandingan Delta antar Task",
        "🔗 Encoding",
    ])

    # ---- Tab 1: Delta ----
    with tab_delta:
        # Panel filter (hanya untuk delta tab)
        with st.expander("Konfigurasi & Filter Batch", expanded=True):
            # --- Select All / Hapus Semua helpers ---
            def _select_all(key, options):
                st.session_state[key] = list(options)
            def _clear_all(key):
                st.session_state[key] = []

            col1, col2, col3 = st.columns(3)

            with col1:
                active_category = "Semua"
                if len(categories) > 1:
                    cat_counts = batch_df.groupby("category")["subject"].nunique()
                    cat_labels = [f"Semua ({batch_df['subject'].nunique()})"]
                    for cat in categories:
                        cat_labels.append(f"{cat} ({cat_counts.get(cat, 0)})")
                    selected_cat_label = st.selectbox("Kategori", cat_labels, key="cat_filter")
                    if not selected_cat_label.startswith("Semua"):
                        active_category = selected_cat_label.split(" (")[0]
                elif categories:
                    active_category = categories[0]

            with col2:
                selected_scenarios = scenarios
                if len(scenarios) > 1:
                    sa1, sa2 = st.columns(2)
                    sa1.button("Semua", key="sa_sc", on_click=_select_all, args=("scenario_filter", scenarios))
                    sa2.button("Hapus", key="cl_sc", on_click=_clear_all, args=("scenario_filter",))
                    selected_scenarios = st.multiselect("Filter Skenario", scenarios, default=[], key="scenario_filter")
                st.markdown("---")
                selected_subbands = all_subbands
                if all_subbands:
                    sb1, sb2 = st.columns(2)
                    sb1.button("Semua", key="sa_sb", on_click=_select_all, args=("batch_subband", all_subbands))
                    sb2.button("Hapus", key="cl_sb", on_click=_clear_all, args=("batch_subband",))
                    selected_subbands = st.multiselect("Filter Subband", all_subbands, default=[], key="batch_subband")

            with col3:
                selected_times = times
                if len(times) > 1:
                    ta1, ta2 = st.columns(2)
                    ta1.button("Semua", key="sa_tm", on_click=_select_all, args=("time_filter", times))
                    ta2.button("Hapus", key="cl_tm", on_click=_clear_all, args=("time_filter",))
                    selected_times = st.multiselect("Filter Time", times, default=[], key="time_filter")
                st.markdown("---")
                selected_channels_batch = all_channels
                if all_channels:
                    ca1, ca2 = st.columns(2)
                    ca1.button("Semua", key="sa_ch", on_click=_select_all, args=("batch_channel", all_channels))
                    ca2.button("Hapus", key="cl_ch", on_click=_clear_all, args=("batch_channel",))
                    selected_channels_batch = st.multiselect("Filter Channel", all_channels, default=[], key="batch_channel")

        # Aplikasikan filter
        filtered_df = batch_df.copy()
        if active_category != "Semua":
            filtered_df = filtered_df[filtered_df["category"] == active_category]
        if selected_scenarios:
            filtered_df = filtered_df[filtered_df["scenario"].isin(selected_scenarios)]
        if selected_times:
            filtered_df = filtered_df[filtered_df["time"].isin(selected_times)]
        if selected_subbands:
            filtered_df = filtered_df[filtered_df["subband"].isin(selected_subbands)]
        if selected_channels_batch:
            filtered_df = filtered_df[filtered_df["channel"].isin(selected_channels_batch)]

        meta_cols = {"filename", "category", "subject", "time", "scenario", "task", "channel", "subband"}
        feat_cols = [c for c in filtered_df.columns if c not in meta_cols]

        # Tab: Analisis Delta + Fitur + ERD/ERS
        _render_delta_tab(
            filtered_df, batch_df, batch_tasks, feat_cols,
            categories, selected_scenarios, selected_times,
        )

        # Tabel Fitur per Task (non-delta)
        _render_feature_per_task_table(filtered_df, batch_tasks, feat_cols)

        # ERD/ERS
        _render_erd_ers(filtered_df, batch_tasks, feat_cols)

        # Connectivity (PLI/wPLI)
        _render_connectivity_batch(filtered_df)

    # ---- Tab 2: Encoding ----
    with tab_encoding:
        render_post_batch_encoding()


# ---------------------------------------------------------------------------
# Tab Analisis Delta
# ---------------------------------------------------------------------------

def _render_delta_tab(filtered_df, batch_df, batch_tasks, feat_cols,
                       categories, selected_scenarios, selected_times):
    """Render analisis delta."""
    if len(batch_tasks) < 2:
        st.info("Diperlukan minimal 2 task untuk analisis delta.")
        return

    st.markdown('<p class="section-title">Perbandingan Delta antar Task</p>', unsafe_allow_html=True)

    col_a, col_b = st.columns(2)
    with col_a:
        task_a = st.selectbox("Task A (Kondisi)", batch_tasks, index=0, key="delta_task_a")
    with col_b:
        default_b = 1 if len(batch_tasks) > 1 else 0
        task_b = st.selectbox("Task B (Baseline)", batch_tasks, index=default_b, key="delta_task_b")

    if task_a == task_b:
        st.warning("Pilih dua task yang berbeda.")
        return

    sel_delta_feat = st.selectbox("Fitur Delta", feat_cols, key="delta_feat_select")

    delta_df, agg_df = DeltaCalculator.calculate_task_delta(filtered_df, task_a, task_b, feat_cols)

    if delta_df.empty:
        st.warning(f"Tidak ada data yang cocok untuk {task_a} vs {task_b}.")
        return

    # Visualisasi checkboxes
    st.markdown("**Pilih Visualisasi:**")
    show_delta_table = st.checkbox("Tabel Delta", value=False, key="show_delta_tbl")
    show_transition = st.checkbox("Transition Delta per Group", value=False, key="show_transition")
    show_delta_scatter = st.checkbox("Scatter Plot per Subjek", value=False, key="show_delta_scatter")
    show_delta_heat = st.checkbox("Heatmap Delta", value=False, key="show_delta_heat")

    dcol = f"delta_{sel_delta_feat}"

    # Tabel
    if show_delta_table:
        _render_delta_tables(delta_df, agg_df, task_a, task_b)

    # Transition
    if show_transition:
        _render_transition_delta(
            filtered_df, task_a, task_b, sel_delta_feat,
            selected_scenarios, selected_times,
        )

    # Scatter
    if show_delta_scatter and dcol in delta_df.columns:
        _render_scatter_plot(delta_df, filtered_df, dcol, sel_delta_feat, task_a, task_b)

    # Heatmap
    if show_delta_heat:
        fig_heat = ComparisonPlots.plot_delta_heatmap(agg_df, sel_delta_feat, task_a, task_b)
        if fig_heat:
            st.plotly_chart(fig_heat, use_container_width=True)


# ---------------------------------------------------------------------------
# Sub-renders untuk delta tab
# ---------------------------------------------------------------------------

def _render_delta_tables(delta_df, agg_df, task_a, task_b):
    """Render tabel delta dan agregat."""
    with st.expander("Tabel Delta Lengkap", expanded=True):
        # Warna per kolom untuk tabel lengkap
        meta_cols = {"filename", "category", "subject", "time", "scenario",
                     "task", "channel", "subband"}

        def _color_full_cols(col):
            if col.name in meta_cols:
                return ["background-color: #e3f2fd"] * len(col)   # biru muda
            elif col.name.endswith(f"_{task_a}"):
                return ["background-color: #e8f5e9"] * len(col)   # hijau muda
            elif col.name.endswith(f"_{task_b}"):
                return ["background-color: #fff3e0"] * len(col)   # oranye muda
            elif col.name.startswith("delta_"):
                return ["background-color: #fce4ec"] * len(col)   # pink muda
            return [""] * len(col)

        styled_full = delta_df.style.apply(_color_full_cols)
        st.dataframe(styled_full, use_container_width=True, height=320, hide_index=True)
        dl1, dl2 = st.columns(2)
        with dl1:
            csv_data = delta_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                "Download CSV", csv_data,
                file_name=f"delta_{task_a}_vs_{task_b}.csv",
                mime="text/csv", key="dl_delta_csv",
            )
        with dl2:
            excel_buf = io.BytesIO()
            with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
                _format_micro_units(delta_df).to_excel(writer, index=False, sheet_name="Delta")
                _format_micro_units(agg_df).to_excel(writer, index=False, sheet_name="Agregat")
            st.download_button(
                "Download Excel", excel_buf.getvalue(),
                file_name=f"delta_{task_a}_vs_{task_b}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_delta_xlsx",
            )

        # --- Fokus satu fitur ---
        st.markdown("---")
        st.markdown("**Fokus Satu Fitur**")
        st.caption("Pilih satu fitur untuk menampilkan tabel yang lebih ringkas.")

        meta_cols = {"filename", "category", "subject", "time", "scenario",
                     "task", "channel", "subband"}
        # Deteksi fitur unik dari kolom delta_*
        delta_feat_cols = [c for c in delta_df.columns
                          if c.startswith("delta_")]
        base_features = [c.replace("delta_", "", 1) for c in delta_feat_cols]

        if base_features:
            sel_focus = st.selectbox(
                "Pilih Fitur", base_features, key="focus_feat_select",
            )

            # Kolom yang relevan: meta + task_a value + task_b value + delta
            col_a = f"{sel_focus}_{task_a}"
            col_b = f"{sel_focus}_{task_b}"
            col_delta = f"delta_{sel_focus}"

            # Kumpulkan kolom meta yang ada di delta_df
            keep_cols = [c for c in delta_df.columns if c in meta_cols]
            # Tambahkan kolom fitur yang ada
            for fc in [col_a, col_b, col_delta]:
                if fc in delta_df.columns:
                    keep_cols.append(fc)
            # Fallback: jika kolom per-task tidak ada, cari kolom asli
            if col_a not in delta_df.columns and col_b not in delta_df.columns:
                if sel_focus in delta_df.columns:
                    keep_cols.append(sel_focus)
                if col_delta in delta_df.columns and col_delta not in keep_cols:
                    keep_cols.append(col_delta)

            focus_df = delta_df[keep_cols].copy()

            # Warna per kolom
            def _color_focus_cols(col):
                if col.name in meta_cols:
                    return ["background-color: #e3f2fd"] * len(col)  # biru muda (meta)
                elif col.name == col_a:
                    return ["background-color: #e8f5e9"] * len(col)  # hijau muda (Task A)
                elif col.name == col_b:
                    return ["background-color: #fff3e0"] * len(col)  # oranye muda (Task B)
                elif col.name == col_delta:
                    return ["background-color: #fce4ec"] * len(col)  # pink muda (Delta)
                return [""] * len(col)

            styled_focus = focus_df.style.apply(_color_focus_cols)
            st.dataframe(styled_focus, use_container_width=True, height=320,
                         hide_index=True)

            fl1, fl2 = st.columns(2)
            with fl1:
                csv_focus = focus_df.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "Download CSV (Fokus)", csv_focus,
                    file_name=f"delta_{sel_focus}_{task_a}_vs_{task_b}.csv",
                    mime="text/csv", key="dl_focus_csv",
                )
            with fl2:
                excel_focus = io.BytesIO()
                with pd.ExcelWriter(excel_focus, engine="openpyxl") as writer:
                    _format_micro_units(focus_df).to_excel(writer, index=False,
                                     sheet_name=sel_focus[:31])
                st.download_button(
                    "Download Excel (Fokus)", excel_focus.getvalue(),
                    file_name=f"delta_{sel_focus}_{task_a}_vs_{task_b}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_focus_xlsx",
                )



def _render_transition_delta(filtered_df, task_a, task_b, sel_delta_feat,
                              selected_scenarios, selected_times):
    """Render transition delta per group."""
    st.markdown("**Transition Delta per Group (Per-Subject)**")
    st.caption("Delta dihitung per pasien dulu, lalu rata-rata per group.")
    transition_df = DeltaCalculator.compute_transition_table(
        filtered_df, task_a, task_b, sel_delta_feat,
        scenarios=selected_scenarios if selected_scenarios else None,
        sessions=selected_times if selected_times else None,
    )
    if not transition_df.empty:
        fig_trans = ComparisonPlots.plot_transition_deltas(
            transition_df, sel_delta_feat,
            title=f"Transition: {task_a} -> {task_b}",
        )
        if fig_trans:
            st.plotly_chart(fig_trans, use_container_width=True)
        st.dataframe(transition_df, use_container_width=True, hide_index=True)


def _render_scatter_plot(delta_df, filtered_df, dcol, sel_delta_feat, task_a, task_b):
    """Render scatter plot delta per subjek."""
    meta_cols_available = ["filename"]
    for mc in ["subject", "category"]:
        if mc in filtered_df.columns:
            meta_cols_available.append(mc)

    meta_map = filtered_df.drop_duplicates(subset=["filename"])[meta_cols_available].copy()

    extra_cols = [c for c in meta_cols_available if c != "filename" and c not in delta_df.columns]
    if extra_cols:
        scatter_df = delta_df.merge(meta_map[["filename"] + extra_cols], on="filename", how="left")
    else:
        scatter_df = delta_df.copy()

    scatter_df["ch_sub"] = scatter_df["channel"] + " / " + scatter_df["subband"]

    has_category = "category" in scatter_df.columns
    has_subject = "subject" in scatter_df.columns

    if has_category:
        scatter_cats = sorted(scatter_df["category"].dropna().unique().tolist())
        if scatter_cats:
            sel_scatter_cat = st.multiselect(
                "Filter Kategori (Scatter)", scatter_cats,
                default=scatter_cats, key="scatter_cat_filter",
            )
            if sel_scatter_cat:
                scatter_df = scatter_df[scatter_df["category"].isin(sel_scatter_cat)]

    fig_scatter = go.Figure()

    if has_subject:
        subjects = sorted(scatter_df["subject"].dropna().unique().tolist())
        for subj in subjects:
            s_data = scatter_df[scatter_df["subject"] == subj]
            cat_label = s_data["category"].iloc[0] if has_category and not s_data.empty else ""
            fig_scatter.add_trace(go.Scatter(
                x=s_data["ch_sub"], y=s_data[dcol],
                mode="markers",
                name=f"{subj} ({cat_label})" if cat_label else subj,
                marker=dict(size=10, opacity=0.8),
            ))
    else:
        filenames = sorted(scatter_df["filename"].dropna().unique().tolist())
        for fname in filenames[:20]:
            s_data = scatter_df[scatter_df["filename"] == fname]
            short_name = fname.replace(".edf", "").split("/")[-1]
            fig_scatter.add_trace(go.Scatter(
                x=s_data["ch_sub"], y=s_data[dcol],
                mode="markers", name=short_name,
                marker=dict(size=10, opacity=0.8),
            ))

    fig_scatter.update_layout(
        title=f"Delta {sel_delta_feat.capitalize()} per Subjek ({task_a} - {task_b})",
        xaxis_title="Channel / Subband",
        yaxis_title=f"Delta {sel_delta_feat}",
        template="plotly_white",
        height=500,
        legend_title="Subjek" if has_subject else "File",
        plot_bgcolor="white",
        paper_bgcolor="white",
    )
    fig_scatter.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5)
    st.plotly_chart(fig_scatter, use_container_width=True)


# ---------------------------------------------------------------------------
# Helper: Excel micro-unit formatting
# ---------------------------------------------------------------------------

def _format_micro_units(df):
    """Konversi kolom numerik yang nilainya sangat kecil ke satuan µ (micro).

    Jika rata-rata absolut kolom < 1e-3, kalikan dengan 1e6 dan tambah (µ)
    pada nama kolom agar lebih mudah dibaca.
    """
    df_out = df.copy()
    rename_map = {}
    for col in df_out.select_dtypes(include=[np.number]).columns:
        mean_abs = df_out[col].abs().mean()
        if 0 < mean_abs < 1e-3:
            df_out[col] = df_out[col] * 1e6
            rename_map[col] = f"{col} (µ)"
    if rename_map:
        df_out.rename(columns=rename_map, inplace=True)
    return df_out


# ---------------------------------------------------------------------------
# Non-delta: Tabel Fitur per Task
# ---------------------------------------------------------------------------

def _render_feature_per_task_table(filtered_df, batch_tasks, feat_cols):
    """Render tabel fitur per task tanpa delta (format pivot)."""
    if not batch_tasks or not feat_cols:
        return

    st.markdown('<p class="section-title">Tabel Fitur per Task</p>', unsafe_allow_html=True)
    st.caption("Data fitur mentah per task, tanpa perhitungan delta.")

    show_feat_table = st.checkbox("Tampilkan Tabel Fitur per Task", value=False, key="show_feat_per_task")
    if not show_feat_table:
        return

    # Tambahkan opsi "Semua Task" dan "Semua Fitur"
    task_opts = ["Semua Task"] + batch_tasks
    feat_opts = ["Semua Fitur"] + feat_cols

    sel_task_table = st.selectbox("Pilih Task", task_opts, key="feat_task_select")
    sel_feat_table = st.selectbox("Pilih Fitur", feat_opts, key="feat_table_feat")

    hide_delta = st.checkbox("Sembunyikan Subband Delta (Tabel Tab Khusus)", value=True, key="hide_delta_chk")

    # Filter berdasarkan Task
    if sel_task_table == "Semua Task":
        task_df = filtered_df.copy()
        sel_task_list = batch_tasks # iterasi semua
    else:
        task_df = filtered_df[filtered_df["task"] == sel_task_table].copy()
        sel_task_list = [sel_task_table]

    if hide_delta and "subband" in task_df.columns:
        task_df = task_df[task_df["subband"] != "Delta"]

    if task_df.empty:
        st.warning(f"Tidak ada data untuk task '{sel_task_table}'.")
        return

    # Filter berdasarkan Fitur
    if sel_feat_table == "Semua Fitur":
        sel_feat_list = feat_cols # iterasi semua
    else:
        sel_feat_list = [sel_feat_table]

    # Sembunyikan tabel di UI jika opsi "Semua" dipilih agar web tidak perlu scroll panjang
    if sel_task_table == "Semua Task" or sel_feat_table == "Semua Fitur":
        st.info("💡 Menampilkan **PREVIEW** (Contoh 1 Task & 1 Fitur pertama) agar antarmuka tidak berat dan scroll terlalu panjang. Seluruh data (semua task/fitur) tetap diekstrak utuh ketika Anda melakukan Download Tabel Excel di bawah.")
        ui_task_list = sel_task_list[:1]
        ui_feat_list = sel_feat_list[:1]
    else:
        ui_task_list = sel_task_list
        ui_feat_list = sel_feat_list

    for current_task in ui_task_list:
        st.markdown(f"### Menganalisis Task: {current_task}")
        curr_task_df = task_df[task_df["task"] == current_task].copy()
        
        if curr_task_df.empty:
            continue
            
        for current_feat in ui_feat_list:
            st.markdown(f"#### Fitur: {current_feat}")
            # Cek apakah nilai sangat kecil (perlu konversi ke micro)
            feat_vals = pd.to_numeric(curr_task_df[current_feat], errors="coerce").dropna()
            use_micro = False
            if not feat_vals.empty and 0 < feat_vals.abs().mean() < 1e-3:
                use_micro = True
                st.info(f"💡 Nilai sangat kecil dideteksi untuk fitur {current_feat}. Data ditampilkan dalam satuan mikro (µ * 10^6).")

            # Pivot: baris = subject, kolom = scenario, group by channel+subband
            has_scenario = "scenario" in curr_task_df.columns
            has_subject = "subject" in curr_task_df.columns
            has_channel = "channel" in curr_task_df.columns
            has_subband = "subband" in curr_task_df.columns

            if not (has_subject and has_channel and has_subband):
                if use_micro:
                    curr_task_df[current_feat] = curr_task_df[current_feat] * 1e6
                st.dataframe(curr_task_df, use_container_width=True, hide_index=True)
                continue

            # Subband color mapping
            subband_colors = {
                "Mu": "#FFF3E0", "Low_Beta": "#E8F5E9", "High_Beta": "#E3F2FD",
                "Alpha": "#FCE4EC", "Beta": "#F3E5F5", "Delta": "#FFFDE7",
                "Theta": "#E0F7FA", "Gamma": "#FBE9E7",
            }

            channels_in_data = sorted(curr_task_df["channel"].unique())
            subbands_in_data = sorted(curr_task_df["subband"].unique())

            # Tambahkan kolom subject_time agar bisa dibedakan per time
            has_time = "time" in curr_task_df.columns
            if has_time:
                curr_task_df["subject_time"] = curr_task_df["subject"] + " (" + curr_task_df["time"].astype(str) + ")"
            else:
                curr_task_df["subject_time"] = curr_task_df["subject"]

            for ch in channels_in_data:
                st.markdown(f"**Channel: {ch}**")
                for sb in subbands_in_data:
                    sb_df = curr_task_df[
                        (curr_task_df["channel"] == ch) & (curr_task_df["subband"] == sb)
                    ].copy()
                    if sb_df.empty:
                        continue

                    if use_micro:
                        sb_df[current_feat] = sb_df[current_feat] * 1e6

                    pivot_index = "subject_time"
                    if has_scenario:
                        try:
                            pivot = sb_df.pivot_table(
                                index=pivot_index, columns="scenario",
                                values=current_feat, aggfunc="first",
                            )
                            pivot = pivot.reindex(sorted(pivot.columns), axis=1)
                        except Exception:
                            pivot = sb_df[[pivot_index, current_feat]].set_index(pivot_index)
                    else:
                        pivot = sb_df[[pivot_index, current_feat]].set_index(pivot_index)
                    pivot.index.name = "subject"

                    # Berikan format desimal khusus
                    disp_feat = f"{current_feat} (µ)" if use_micro else current_feat
                    bg_color = subband_colors.get(sb, "#FFFFFF")
                    styled = pivot.style.set_properties(**{"background-color": bg_color}).format("{:.6f}")
                    
                    st.markdown(f"*{sb}* — {disp_feat}")
                    st.dataframe(styled, use_container_width=True, height=min(35 * (len(pivot) + 1), 400))
                st.markdown("---")

    # Download Excel (Custom Layout)
    import io
    from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
    
    # Inisialisasi Border
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    excel_buf = io.BytesIO()
    with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
        has_cat = "category" in task_df.columns
        categories = sorted(task_df["category"].dropna().astype(str).unique().tolist()) if has_cat else ["Semua"]
        if not categories:
            categories = ["Semua"]

        ch_colors = {
            0: "FFE699", 1: "9BC2E6", 2: "C6E0B4",
            3: "F4B084", 4: "D9D9D9", 5: "B4A7D6",
        }

        for cat in categories:
            if cat == "Semua":
                df_cat = task_df.copy()
            else:
                df_cat = task_df[task_df["category"] == cat].copy()

            if df_cat.empty:
                continue
                
            for current_task in sel_task_list:
                df_cat_task = df_cat[df_cat["task"] == current_task].copy()
                if df_cat_task.empty:
                    continue
                    
                for current_feat in sel_feat_list:
                    df_cat_task_feat = df_cat_task.copy()

                    import re
                    safe_feat = re.sub(r'[\\/\*\?:\\[\\]]', '', str(current_feat))
                    safe_task = re.sub(r'[\\/\*\?:\\[\\]]', '', str(current_task))
                    safe_cat = re.sub(r'[\\/\*\?:\\[\\]]', '', str(cat))

                    sheet_name = f"{safe_feat}_{safe_task}_{safe_cat}"[:31]

                    # Buat sheet via pandas dict biar aman
                    pd.DataFrame().to_excel(writer, sheet_name=sheet_name)
                    ws = writer.sheets[sheet_name]

                    # Hapus konten dari cell awal pandas (index & blank)
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.value = None

                    ws.cell(row=1, column=1, value=current_task).font = Font(bold=True)

                    channels_arr = sorted(df_cat_task_feat["channel"].unique().tolist()) if "channel" in df_cat_task_feat.columns else []
                    subbands_arr = sorted(df_cat_task_feat["subband"].unique().tolist()) if "subband" in df_cat_task_feat.columns else []
                    # Gabungkan subject+time agar Excel juga menampilkan info time
                    has_time_col = "time" in df_cat_task_feat.columns
                    if has_time_col:
                        df_cat_task_feat["subject_time"] = df_cat_task_feat["subject"] + " (" + df_cat_task_feat["time"].astype(str) + ")"
                    else:
                        df_cat_task_feat["subject_time"] = df_cat_task_feat["subject"]
                    subjects_arr = sorted(df_cat_task_feat["subject_time"].unique().tolist()) if "subject_time" in df_cat_task_feat.columns else []

                    has_scen = "scenario" in df_cat_task_feat.columns
                    scenarios_arr = sorted(df_cat_task_feat["scenario"].dropna().unique().tolist()) if has_scen else ["Value"]

                    if not subjects_arr or not channels_arr or not subbands_arr:
                        continue

                    # Cek apakai pakai unit micro (µ)
                    feat_vals = pd.to_numeric(df_cat_task_feat[current_feat], errors="coerce").dropna()
                    use_micro = False
                    if not feat_vals.empty:
                        if 0 < feat_vals.abs().mean() < 1e-3:
                            use_micro = True
                    disp_feat_name = f"{current_feat} (µ)" if use_micro else current_feat

                    # Mapping nilai
                    val_map = {}
                    for _, r in df_cat_task_feat.iterrows():
                        val = r.get(current_feat)
                        if pd.isna(val):
                            continue
                        try:
                            val = float(val)
                            if use_micro:
                                val *= 1e6
                        except ValueError:
                            pass
                        ch = r.get("channel")
                        sb = r.get("subband")
                        subj = r.get("subject_time")
                        scen = r.get("scenario") if has_scen else "Value"
                        val_map[(ch, sb, subj, scen)] = val

                    for c_idx, ch in enumerate(channels_arr):
                        grid_row = c_idx // 2
                        grid_col = c_idx % 2

                        # Menentukan posisi baris dan kolom untuk channel ini
                        # Tiap subband ukurannya: 2 header + len(subjects_arr) data + 2 row jarak = len(subjects_arr) + 4
                        row_step_per_subband = len(subjects_arr) + 4
                        start_r = 3 + grid_row * (len(subbands_arr) * row_step_per_subband)
                        start_c = 1 + grid_col * (len(scenarios_arr) + 3)

                        ch_color = PatternFill("solid", fgColor=ch_colors.get(c_idx % 6, "FFFFFF"))
                        sb_color = PatternFill("solid", fgColor="F4B084") # orange

                        # Menghitung sampai mana cell channel akan digabung (batas akhir adalah isi data subjekt terakhir)
                        end_r_for_ch = start_r + len(subbands_arr) * row_step_per_subband - 3

                        c_cell = ws.cell(row=start_r, column=start_c, value=ch)
                        c_cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
                        c_cell.fill = ch_color
                        c_cell.font = Font(bold=True)
                        c_cell.border = thin_border
                        
                        if end_r_for_ch > start_r:
                            ws.merge_cells(start_row=start_r, start_column=start_c, end_row=end_r_for_ch, end_column=start_c)
                            # Set border untuk cell yang dimerge
                            for row in ws.iter_rows(min_row=start_r, max_row=end_r_for_ch, min_col=start_c, max_col=start_c):
                                for cell in row:
                                    cell.border = thin_border

                        curr_r = start_r
                        for sb in subbands_arr:
                            # Header: Subband name
                            sb_cell = ws.cell(row=curr_r, column=start_c+1, value=sb)
                            sb_cell.fill = sb_color
                            sb_cell.font = Font(bold=True)
                            sb_cell.border = thin_border

                            # Header: Feature name (spanning scenarios)
                            f_cell = ws.cell(row=curr_r, column=start_c+2, value=disp_feat_name)
                            f_cell.font = Font(bold=True)
                            f_cell.alignment = Alignment(horizontal="center")
                            f_cell.border = thin_border
                            
                            if len(scenarios_arr) > 1:
                                ws.merge_cells(start_row=curr_r, start_column=start_c+2, end_row=curr_r, end_column=start_c+1+len(scenarios_arr))
                                # Set border untuk cell yang dimerge
                                for row in ws.iter_rows(min_row=curr_r, max_row=curr_r, min_col=start_c+2, max_col=start_c+1+len(scenarios_arr)):
                                    for cell in row:
                                        cell.border = thin_border

                            curr_r += 1
                            # Header: "ID" and scenarios names
                            id_cell = ws.cell(row=curr_r, column=start_c+1, value="ID")
                            id_cell.font = Font(bold=True)
                            id_cell.border = thin_border
                            
                            for s_idx, scen in enumerate(scenarios_arr):
                                sc_cell = ws.cell(row=curr_r, column=start_c+2+s_idx, value=scen)
                                sc_cell.font = Font(bold=True)
                                sc_cell.alignment = Alignment(horizontal="center")
                                sc_cell.border = thin_border
                            
                            curr_r += 1
                            # Data rows
                            for subj in subjects_arr:
                                # Row header: Subject ID
                                subj_cell = ws.cell(row=curr_r, column=start_c+1, value=subj)
                                subj_cell.border = thin_border
                                
                                # Data values per scenario
                                for s_idx, scen in enumerate(scenarios_arr):
                                    v = val_map.get((ch, sb, subj, scen))
                                    v_cell = ws.cell(row=curr_r, column=start_c+2+s_idx, value=v)
                                    v_cell.border = thin_border
                                curr_r += 1

                            # Spacer before next subband
                            curr_r += 2

    st.download_button(
        label="📥 Download Tabel Excel (Grid Custom + Multiple Fitur/Task)",
        data=excel_buf.getvalue(),
        file_name="batch_features_grid_all.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="btn_dl_feat_all"
    )




# ---------------------------------------------------------------------------
# ERD/ERS Analysis
# ---------------------------------------------------------------------------

def _render_erd_ers(filtered_df, batch_tasks, feat_cols):
    """Render analisis ERD/ERS (paired: Resting → Task)."""
    if "Resting" not in batch_tasks or len(batch_tasks) < 2:
        return

    st.markdown('<p class="section-title">ERD/ERS Analysis (Paired Baseline)</p>', unsafe_allow_html=True)
    st.caption(
        "Event-Related Desynchronization (ERD) / Synchronization (ERS). "
        "Baseline = Resting yang muncul tepat sebelum Task. "
        "Negatif = ERD (penurunan power), Positif = ERS (kenaikan power)."
    )

    show_erd = st.checkbox("Tampilkan ERD/ERS", value=False, key="show_erd_ers")
    if not show_erd:
        return

    # Ambil pre-computed ERD/ERS paired dari session state
    erd_df = st.session_state.get("batch_erd_df", pd.DataFrame())

    if erd_df is None or erd_df.empty:
        st.warning("Tidak ada data ERD/ERS (pastikan ada pasangan Resting → Task dalam annotations).")
        return

    # Terapkan filter yang sama dari filtered_df
    if "filename" in filtered_df.columns and "filename" in erd_df.columns:
        valid_files = filtered_df["filename"].unique()
        erd_df = erd_df[erd_df["filename"].isin(valid_files)].copy()
    if "channel" in filtered_df.columns and "channel" in erd_df.columns:
        valid_channels = filtered_df["channel"].unique()
        erd_df = erd_df[erd_df["channel"].isin(valid_channels)].copy()
    if "subband" in filtered_df.columns and "subband" in erd_df.columns:
        valid_subbands = filtered_df["subband"].unique()
        erd_df = erd_df[erd_df["subband"].isin(valid_subbands)].copy()

    if erd_df.empty:
        st.warning("Tidak ada data ERD/ERS setelah filter diterapkan.")
        return

    # --- Pemilihan Baseline & Task ---
    available_tasks = sorted(erd_df["task"].unique().tolist()) if "task" in erd_df.columns else []
    if not available_tasks:
        st.warning("Tidak ada task ERD/ERS yang tersedia.")
        return

    col_bl, col_tk = st.columns(2)
    with col_bl:
        st.markdown("**Baseline**")
        st.info("Resting (otomatis)")
    with col_tk:
        sel_task = st.selectbox("Task (kondisi)", available_tasks, key="erd_task_sel")

    # Filter berdasarkan task yang dipilih
    erd_df = erd_df[erd_df["task"] == sel_task].copy()
    if erd_df.empty:
        st.warning(f"Tidak ada data ERD/ERS untuk task '{sel_task}'.")
        return

    # Susun tabel tampilan — kolom sudah fixed: baseline_power, task_power, erd_ers_pct
    meta_display = ["filename", "category", "subject", "time", "scenario", "task",
                    "channel", "subband"]
    display_cols = [c for c in meta_display if c in erd_df.columns]
    for c in ["baseline_power", "task_power", "erd_ers_pct"]:
        if c in erd_df.columns:
            display_cols.append(c)

    out_erd = erd_df[display_cols].copy()

    # Warna: negatif = merah (ERD), positif = hijau (ERS)
    def _color_erd(val):
        if isinstance(val, (int, float)):
            if val < 0:
                return "background-color: #FFCDD2"  # merah muda
            elif val > 0:
                return "background-color: #C8E6C9"  # hijau muda
        return ""

    styled_erd = out_erd.style.map(_color_erd, subset=["erd_ers_pct"])
    st.dataframe(styled_erd, use_container_width=True, height=400, hide_index=True)

    # Download
    fl1, fl2 = st.columns(2)
    with fl1:
        csv_erd = out_erd.to_csv(index=False).encode("utf-8")
        st.download_button(
            "Download CSV (ERD/ERS)", csv_erd,
            file_name="erd_ers_band_power.csv",
            mime="text/csv", key="dl_erd_csv",
        )
    with fl2:
        excel_erd = io.BytesIO()
        with pd.ExcelWriter(excel_erd, engine="openpyxl") as writer:
            out_excel = _format_micro_units(out_erd)
            out_excel.to_excel(writer, index=False, sheet_name="ERD_ERS")
        st.download_button(
            "Download Excel (ERD/ERS)", excel_erd.getvalue(),
            file_name="erd_ers_band_power.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_erd_xlsx",
        )


# ---------------------------------------------------------------------------
# Connectivity Analysis (Batch)
# ---------------------------------------------------------------------------

def _render_connectivity_batch(filtered_df):
    """Render tabel Connectivity PLI/wPLI batch."""
    conn_df = st.session_state.get("batch_conn_df", pd.DataFrame())
    
    if conn_df is None or conn_df.empty:
        return
        
    st.markdown('<p class="section-title">Connectivity (PLI / wPLI)</p>', unsafe_allow_html=True)
    
    # Filter sesuai dengan filtered_df
    if "filename" in filtered_df.columns and "filename" in conn_df.columns:
        valid_files = filtered_df["filename"].unique()
        conn_df = conn_df[conn_df["filename"].isin(valid_files)].copy()
    if "subband" in filtered_df.columns and "subband" in conn_df.columns:
        valid_subbands = filtered_df["subband"].unique()
        conn_df = conn_df[conn_df["subband"].isin(valid_subbands)].copy()
        
    if conn_df.empty:
        st.warning("Tidak ada data Connectivity setelah filter diterapkan.")
        return
        
    st.dataframe(conn_df, use_container_width=True, height=400, hide_index=True)
    
    # Download
    fl1, fl2 = st.columns(2)
    with fl1:
        csv_conn = conn_df.to_csv(index=False).encode("utf-8")
        method_name = conn_df["method"].iloc[0] if "method" in conn_df.columns and not conn_df.empty else "connectivity"
        st.download_button(
            f"Download CSV ({method_name.upper()})", csv_conn,
            file_name=f"{method_name}_results.csv",
            mime="text/csv", key="dl_conn_csv",
        )
    with fl2:
        excel_conn = io.BytesIO()
        with pd.ExcelWriter(excel_conn, engine="openpyxl") as writer:
            conn_df.to_excel(writer, index=False, sheet_name="Connectivity")
        st.download_button(
            f"Download Excel ({method_name.upper()})", excel_conn.getvalue(),
            file_name=f"{method_name}_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_conn_xlsx",
        )
